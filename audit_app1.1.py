import streamlit as st
import pandas as pd
import os
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# --- 1. ПЕРВИЧНАЯ НАСТРОЙКА ---
st.set_page_config(page_title="Аудит ИТ и ИБ 2026", layout="wide", page_icon="🛡️")

# --- 2. БРЕНДИНГ И ЛОГОТИП ---
if os.path.exists("logo.png"):
    st.image("logo.png", width=300)
else:
    st.title("Ivan Rudoy | IT Audit & Consulting")

st.markdown("### 📞 **Хотите такой опросник — звоните!**")
st.divider()

st.title("📋 Опросник: Технический аудит ИТ и ИБ (2026)")

# Инициализация переменных для данных и скоринга
data = {}
score = 0

# Вспомогательные функции для выбора
def get_choice(label, options, key):
    res = st.selectbox(label, options + ["Другое"], key=f"sel_{key}")
    return st.text_input(f"Укажите свой вариант ({label}):", key=f"oth_{key}") if res == "Другое" else res

def get_multi(label, options, key):
    res = st.multiselect(label, options + ["Другое"], key=f"msel_{key}")
    if "Другое" in res:
        other = st.text_input(f"Дополнительно ({label}):", key=f"moth_{key}")
        return ", ".join([x for x in res if x != "Другое"] + ([other] if other else []))
    return ", ".join(res)

# --- БЛОК 1: ОБЩАЯ ИНФОРМАЦИЯ ---
st.header("Блок 1: Общая информация")
data['1. Сотрудников в штате'] = st.number_input("1. Общее количество сотрудников:", min_value=0, step=1)
has_it = st.toggle("2. В компании есть выделенный ИТ-департамент?")

if has_it:
    col1, col2 = st.columns(2)
    with col1:
        data['1.1. Количество АРМ'] = st.number_input("1.1. Количество конечных точек (АРМ):", min_value=0)
        data['1.2. Физические серверы'] = st.number_input("1.2. Кол-во физических серверов:", min_value=0)
        data['1.3. ОС (*Nix)'] = st.number_input("1.3. Количество серверов на *Nix:", min_value=0)
        data['1.4. Виртуализация'] = get_multi("1.4. Среда виртуализации:", ["VMware", "Hyper-V", "Proxmox"], "virt")
    with col2:
        data['1.2. Виртуальные серверы'] = st.number_input("1.2. Кол-во виртуальных серверов:", min_value=0)
        data['1.3. ОС (Windows)'] = st.number_input("1.3. Количество серверов на Windows:", min_value=0)
        data['1.5. Тикет-система'] = st.text_input("1.5. Используемая Тикет-система:")
        data['1.6. Мониторинг'] = st.text_input("1.6. Система мониторинга:")
    data['1.7. Почтовая система'] = get_choice("1.7. Почта:", ["Cloud", "On-Prem", "Hybrid"], "mail")

# --- БЛОК 2: СЕТЕВАЯ ИНФРАСТРУКТУРА ---
st.header("Блок 2: Сетевая инфраструктура и Интернет")
if st.toggle("Есть собственная сетевая инфраструктура?"):
    data['2.1. Тип канала'] = get_multi("2.1. Тип Интернет-канала:", ["Оптика", "Радиорелейная", "Спутник", "4G/5G"], "net_c")
    data['2.1. Скорость (Мбит/с)'] = st.number_input("Заявленная скорость канала:", min_value=0)
    data['2.2. Core (Ядро)'] = st.text_input("Вендор/модель ядра сети:")
    
    c_net1, c_net2 = st.columns(2)
    with c_net1:
        data['2.3. Коммутаторы L2'] = st.number_input("Количество L2 коммутаторов:", min_value=0)
        data['2.3. Технологии'] = get_multi("Используемые технологии:", ["VLAN", "STP", "BGP", "SD-WAN"], "tech")
    with c_net2:
        data['2.3. Коммутаторы L3'] = st.number_input("Количество L3 коммутаторов:", min_value=0)
        data['2.4. NGFW (Вендор)'] = st.text_input("Межсетевой экран (NGFW):")
        if data['2.4. NGFW (Вендор)']: score += 20 # Скоринг за NGFW

    if st.checkbox("2.5. Наличие Wi-Fi"):
        data['2.5.1. Контроллер'] = st.text_input("Контроллер Wi-Fi:")
        data['2.5.2. Точек доступа'] = st.number_input("Кол-во точек доступа:", min_value=0)

# --- БЛОК 3: ИНФОРМАЦИОННАЯ БЕЗОПАСНОСТЬ (ОСНОВНОЙ СКОРИНГ) ---
st.header("Блок 3: Информационная Безопасность")
if st.toggle("Внедрены системы ИБ?"):
    ib_systems = {
        "DLP (Защита от утечек)": 15,
        "PAM (Контроль доступа)": 10,
        "SIEM/SOC (Мониторинг)": 20,
        "WAF (Защита сайтов)": 10,
        "EDR/Antimalware": 15,
        "Резервное копирование": 20
    }
    for label, points in ib_systems.items():
        if st.checkbox(label):
            data[label] = "Да"
            score += points
        else:
            data[label] = "Нет"
    
    auth = st.radio("3.4. Тип аутентификации:", ["Только пароли", "MFA/2FA внедрена"])
    if auth == "MFA/2FA внедрена":
        score += 10
        data['3.4. Аутентификация'] = "MFA"
    else:
        data['3.4. Аутентификация'] = "Пароли"

# --- БЛОК 4: WEB-РЕСУРСЫ ---
st.header("Блок 4: Web-ресурсы")
if st.toggle("Есть внешние веб-ресурсы?"):
    data['4.1. Хостинг'] = get_choice("4.1. Размещение:", ["Собственный ЦОД", "Облако (локальное)", "Облако (глобальное)"], "hst")
    data['4.2. CMS'] = get_choice("4.2. Платформа:", ["Bitrix", "WordPress", "Custom"], "cms")
    data['4.3. Базы данных'] = get_multi("4.3. СУБД:", ["PostgreSQL", "MySQL", "Oracle", "MS SQL"], "db")

# --- БЛОК 5: РАЗРАБОТКА ---
st.header("Блок 5: Разработка")
if st.toggle("Ведется собственная разработка?"):
    data['5.1. Кол-во разработчиков'] = st.number_input("Разработчиков в штате:", min_value=0)
    data['5.2. Стек'] = get_multi("Основной стек:", ["Java", "Python", ".NET", "PHP", "JS"], "stk")
    data['5.4. Контейнеризация'] = st.text_input("Используемые технологии (Docker/K8s):")

# --- ГЕНЕРАЦИЯ ЭКСПЕРТНОГО EXCEL ---
def create_report(results_dict, total_score):
    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Аналитический отчет"

    # Стили
    blue_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Шапка
    ws.merge_cells('A1:C1')
    ws['A1'] = "ЭКСПЕРТНЫЙ ОТЧЕТ ПО ТЕХНИЧЕСКОМУ АУДИТУ"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Скоринг
    ws['A3'] = "ИНДЕКС ТЕХНИЧЕСКОЙ ЗРЕЛОСТИ:"
    ws['B3'] = f"{total_score} / 100"
    color = "00B050" if total_score > 70 else "FFCC00" if total_score > 40 else "FF0000"
    ws['B3'].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    ws['B3'].font = Font(bold=True)

    # Таблица данных
    headers = ["Параметр", "Ответ клиента", "Аналитика / Риск"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=5, column=i, value=h)
        cell.fill = blue_fill
        cell.font = white_font

    curr_row = 6
    for k, v in results_dict.items():
        ws.cell(row=curr_row, column=1, value=k).border = border
        ws.cell(row=curr_row, column=2, value=str(v)).border = border
        
        # Аналитическая колонка
        if v == "Нет" or v == 0:
            cell_risk = ws.cell(row=curr_row, column=3, value="ТРЕБУЕТ ВНИМАНИЯ")
            cell_risk.font = Font(color="FF0000", bold=True)
        else:
            ws.cell(row=curr_row, column=3, value="В норме")
        
        ws.cell(row=curr_row, column=3).border = border
        curr_row += 1

    # Ширина колонок
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 25

    wb.save(output)
    return output.getvalue()

# --- ВЫГРУЗКА ---
st.divider()
if st.button("🚀 Сформировать экспертную аналитику"):
    if not data:
        st.error("Сначала заполните данные в блоках!")
    else:
        final_excel = create_report(data, min(score, 100))
        st.success(f"Анализ завершен. Индекс зрелости ИБ: {min(score, 100)}/100")
        st.download_button("📥 Скачать экспертный Excel", final_excel, "Audit_Expert_Report.xlsx")

st.info("Разработано Ivan Rudoy. По вопросам автоматизации аудита обращайтесь по телефону.")
